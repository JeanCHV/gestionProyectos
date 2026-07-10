from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from typing import Any
import unicodedata

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel


BASE_DIR = Path(__file__).resolve().parent
TEMPLATE_PATH = BASE_DIR / "static" / "files" / "template_riesgos.xlsx"


def _normalize_date(value: Any) -> str | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, str):
        raw = value.strip()
        if not raw:
            return None
        patterns = [
            "%Y-%m-%d",
            "%d/%m/%Y",
            "%d/%m/%y",
            "%m/%d/%Y",
            "%m/%d/%y",
            "%Y/%m/%d",
        ]
        for pattern in patterns:
            try:
                return datetime.strptime(raw, pattern).date().isoformat()
            except ValueError:
                continue
        return None
    return str(value)


def _normalize_date_with_status(value: Any) -> tuple[str | None, bool]:
    normalized = _normalize_date(value)
    if value in (None, ""):
        return None, True
    if normalized is None:
        return None, False
    if isinstance(value, str):
        raw = value.strip()
        patterns = [
            "%Y-%m-%d",
            "%d/%m/%Y",
            "%d/%m/%y",
            "%m/%d/%Y",
            "%m/%d/%y",
            "%Y/%m/%d",
        ]
        for pattern in patterns:
            try:
                parsed = datetime.strptime(raw, pattern).date().isoformat()
                return parsed, True
            except ValueError:
                continue
        return None, False
    return normalized, True


def _cell_value(cell, wb):
    value = cell.value
    if value is None:
        return None
    if getattr(cell, "is_date", False):
        return value
    if isinstance(value, (int, float)) and cell.number_format and "yy" in str(cell.number_format).lower():
        try:
            return from_excel(value, wb.epoch)
        except Exception:
            return value
    return value


def _sheet_rows(ws, wb, start_row: int = 2) -> list[list[Any]]:
    rows: list[list[Any]] = []
    for row in ws.iter_rows(min_row=start_row, values_only=False):
        values = [_cell_value(cell, wb) for cell in row]
        if any(value not in (None, "") for value in values):
            rows.append(values)
    return rows


def _normalize_header(value: Any) -> str:
    if value in (None, ""):
        return ""
    text = str(value).strip().lower()
    text = unicodedata.normalize("NFKD", text)
    return "".join(char for char in text if not unicodedata.combining(char))


def _normalize_lookup(value: Any) -> str:
    if value in (None, ""):
        return ""
    text = str(value).strip().lower()
    text = unicodedata.normalize("NFKD", text)
    return "".join(char for char in text if not unicodedata.combining(char) and char.isalnum())


def _row_map(headers: list[Any], row: list[Any]) -> dict[str, Any]:
    mapped: dict[str, Any] = {}
    for index, header in enumerate(headers):
        key = _normalize_header(header)
        if key:
            mapped[key] = row[index] if index < len(row) else None
    return mapped


def _pick(mapped: dict[str, Any], *aliases: str, default: Any = None) -> Any:
    for alias in aliases:
        key = _normalize_header(alias)
        if key in mapped and mapped[key] not in (None, ""):
            return mapped[key]
    return default


@dataclass
class ImportedWorkbook:
    project: dict[str, Any] | None
    roles: list[dict[str, Any]]
    deliverables: list[dict[str, Any]]
    threats: list[dict[str, Any]]
    safeguards: list[dict[str, Any]]
    assets: list[dict[str, Any]]
    risks: list[dict[str, Any]]
    mitigations: list[dict[str, Any]]
    warnings: list[str]


def load_template_bytes() -> bytes:
    return TEMPLATE_PATH.read_bytes()


def parse_imported_workbook(file_bytes: bytes) -> ImportedWorkbook:
    wb = load_workbook(BytesIO(file_bytes), data_only=True)

    project_data = None
    warnings: list[str] = []
    if "Proyecto" in wb.sheetnames:
        ws = wb["Proyecto"]
        values = _sheet_rows(ws, wb, start_row=2)
        headers = values[0] if values else []
        if len(values) >= 2:
            row = values[1]
            mapped = _row_map(headers, row)
            raw_start = _pick(mapped, "Fecha Inicio", "Inicio")
            raw_end = _pick(mapped, "Fecha Fin", "Fin")
            start_date, start_ok = _normalize_date_with_status(raw_start)
            end_date, end_ok = _normalize_date_with_status(raw_end)
            if not start_ok and raw_start not in (None, ""):
                warnings.append("Fecha Inicio del proyecto omitida por formato invalido.")
            if not end_ok and raw_end not in (None, ""):
                warnings.append("Fecha Fin del proyecto omitida por formato invalido.")
            project_data = {
                "name": _pick(mapped, "Nombre del proyecto", "Proyecto", "Nombre") or "Proyecto sin nombre",
                "project_type": _pick(mapped, "Tipo", "Tipo de proyecto") or "",
                "company": _pick(mapped, "Empresa", "Compañia", "Compañía") or "",
                "start_date": start_date,
                "end_date": end_date,
                "status": _pick(mapped, "Estado", "Status") or "Activo",
            }

    roles: list[dict[str, Any]] = []
    if "Roles" in wb.sheetnames:
        ws = wb["Roles"]
        rows = _sheet_rows(ws, wb, start_row=2)
        headers = rows[0] if rows else []
        for row in rows[1:]:
            item = _row_map(headers, row)
            role_name = _pick(item, "Rol", "Role") or ""
            if not role_name or _normalize_header(role_name) == _normalize_header("Rol"):
                continue
            roles.append(
                {
                    "role": role_name,
                    "acquisition_type": _pick(item, "Tipo de adquisicion", "Tipo de adquisición", "Adquisicion") or "",
                    "risk_participation": _pick(item, "Participacion en riesgos", "Participación en riesgos") or "",
                }
            )

    deliverables: list[dict[str, Any]] = []
    if "Entregables" in wb.sheetnames:
        ws = wb["Entregables"]
        rows = _sheet_rows(ws, wb, start_row=2)
        headers = rows[0] if rows else []
        for row in rows[1:]:
            item = _row_map(headers, row)
            deliverable_name = _pick(item, "Entregable") or ""
            if not deliverable_name:
                continue
            deliverables.append(
                {
                    "deliverable": deliverable_name,
                    "description": _pick(item, "Descripcion", "Descripción") or "",
                    "main_responsible": _pick(item, "Responsable principal", "Responsable") or "",
                    "status": _pick(item, "Estado", "Status") or "Pendiente",
                }
            )

    threats: list[dict[str, Any]] = []
    if "Amenazas" in wb.sheetnames:
        ws = wb["Amenazas"]
        rows = _sheet_rows(ws, wb, start_row=2)
        headers = rows[0] if rows else []
        for row in rows[1:]:
            item = _row_map(headers, row)
            threat_name = _pick(item, "Amenaza", "Threat") or ""
            if not threat_name:
                continue
            threats.append(
                {
                    "code": _pick(item, "Codigo", "Código", "Code") or "",
                    "threat": threat_name,
                    "affected_asset": _pick(item, "Activo afectado", "Activo", "Activos afectados") or "",
                    "example": _pick(item, "Ejemplo en el proyecto", "Ejemplo") or "",
                }
            )

    safeguards: list[dict[str, Any]] = []
    if "Salvaguardas" in wb.sheetnames:
        ws = wb["Salvaguardas"]
        rows = _sheet_rows(ws, wb, start_row=2)
        headers = rows[0] if rows else []
        for row in rows[1:]:
            item = _row_map(headers, row)
            safeguard_name = _pick(item, "Salvaguarda propuesta", "Salvaguarda", "Safeguard") or ""
            if not safeguard_name:
                continue
            safeguards.append(
                {
                    "threat_name": _pick(item, "Amenaza", "Threat") or "",
                    "safeguard": safeguard_name,
                }
            )

    assets: list[dict[str, Any]] = []
    if "Activos" in wb.sheetnames:
        ws = wb["Activos"]
        rows = _sheet_rows(ws, wb, start_row=2)
        headers = rows[0] if rows else []
        for row in rows[1:]:
            item = _row_map(headers, row)
            assets.append(
                {
                    "name": _pick(item, "Activo", "Nombre del activo", "Nombre") or "",
                    "type": _pick(item, "Tipo de activo", "Tipo") or "",
                    "owner": _pick(item, "Responsable") or "",
                    "value": _pick(item, "Valor para el proyecto", "Valor del activo", "Valor") or "",
                    "status": _pick(item, "Estado", "Status") or "Activo",
                }
            )

    risks: list[dict[str, Any]] = []
    if "Riesgos" in wb.sheetnames:
        ws = wb["Riesgos"]
        rows = _sheet_rows(ws, wb, start_row=2)
        headers = rows[0] if rows else []
        for row in rows[1:]:
            item = _row_map(headers, row)
            risks.append(
                {
                    "name": _pick(item, "Riesgo") or "",
                    "asset_name": _pick(item, "Activo") or "",
                    "description": _pick(item, "Descripcion", "Descripción") or "",
                    "cause": _pick(item, "Causa") or "",
                    "consequence": _pick(item, "Consecuencia") or "",
                    "probability": _pick(item, "Probabilidad") or 0,
                    "impact": _pick(item, "Impacto") or 0,
                    "level": _pick(item, "Nivel") or "",
                    "horizon": _pick(item, "Horizonte") or "",
                    "owner": _pick(item, "Responsable") or "",
                    "status": _pick(item, "Estado", "Status") or "Identificado",
                    "strategy": _pick(item, "Estrategia") or "Mitigar",
                }
            )

    asset_names = {_normalize_lookup(asset["name"]) for asset in assets if asset.get("name")}
    for risk in risks:
        asset_name = risk.get("asset_name") or ""
        if asset_name and _normalize_lookup(asset_name) not in asset_names:
            warnings.append(
                f"Riesgo '{risk.get('name') or ''}' referencia el activo '{asset_name}', pero ese activo no existe en la hoja Activos."
            )

    mitigations: list[dict[str, Any]] = []
    if "Mitigacion" in wb.sheetnames:
        ws = wb["Mitigacion"]
        rows = _sheet_rows(ws, wb, start_row=2)
        headers = rows[0] if rows else []
        for row in rows[1:]:
            item = _row_map(headers, row)
            start_date, start_ok = _normalize_date_with_status(_pick(item, "Inicio"))
            end_date, end_ok = _normalize_date_with_status(_pick(item, "Fin"))
            if not start_ok and _pick(item, "Inicio") not in (None, ""):
                warnings.append(f"Mitigacion '{_pick(item, 'Riesgo') or ''}' con Inicio invalido omitido.")
            if not end_ok and _pick(item, "Fin") not in (None, ""):
                warnings.append(f"Mitigacion '{_pick(item, 'Riesgo') or ''}' con Fin invalido omitido.")
            mitigations.append(
                {
                    "risk_name": _pick(item, "Riesgo") or "",
                    "preventive_action": _pick(item, "Accion preventiva", "Accion Preventiva") or "",
                    "corrective_action": _pick(
                        item,
                        "Accion correctiva / contingencia",
                        "Accion correctiva",
                        "Accion Correctiva",
                        "Contingencia",
                    )
                    or "",
                    "trigger": _pick(item, "Disparador") or "",
                    "owner": _pick(item, "Responsable") or "",
                    "start_date": start_date,
                    "end_date": end_date,
                    "resources": _pick(item, "Recursos") or "",
                    "status": _pick(item, "Estado", "Status") or "Planificado",
                    "evidence": _pick(item, "Evidencia") or "",
                    "strategy": _pick(item, "Estrategia") or "Mitigar",
                    "progress": _pick(item, "Avance") or 0,
                }
            )

    return ImportedWorkbook(
        project=project_data,
        roles=roles,
        deliverables=deliverables,
        threats=threats,
        safeguards=safeguards,
        assets=assets,
        risks=risks,
        mitigations=mitigations,
        warnings=warnings,
    )
